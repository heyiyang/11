#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2021/6/11 21:40 
# @Author : Lemon_Tricy
# @QQ: 2378807189
# Copyright：湖南省零檬信息技术有限公司
"""
接口自动化测试：
1、测试数据从excel用例里的读取出来--read_date（）  --Done
3、发送接口请求 ，得到响应消息-执行结果 --- api_request  --Done
4、执行结果 vs 预期结果对比  -- 同-通过；不同-不通过 --断言 ==结论-pass  fail --Done
5、把最终的结果写入excel表格中 --write_result --Done

第三方库：操作excel表格数据：openpyxl
1、安装: pip / pycharm
2、导入 --
import :导入所有
from xxx import xxx : 导入部分

主要的操作：
1、加载工作簿对象：
2、表单对象：sheet
3、单元格：行列-cell
4、写入数据到单元格  --一定保存才会生效
注意：文件一定要关闭才可以保存成功，否则会报错！
wb = load_workbook("testcase_api_wuye.xlsx")  # 加载工作簿
sheet = wb["register"]  # 选择表单
cell = sheet.cell(row=1,column=1) # 单元格
print(cell.value)
cell.value = "用例编号"  # 重新赋值  --写入数据
print(cell.value)
wb.save("testcase_api_wuye.xlsx")  # 保存文件

把一个功能封装成函数步骤：
1、def  函数名
2、参数化-- 变量值
3、是否需要定义返回值 --- 有没有东西给别人使用
"""
import requests
from openpyxl import load_workbook
#接口请求函数
def api_request(url,data):
    head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    response = requests.post(url=url,json=data,headers=head)   # 方法传参  -关键字
    return  response.json()

# 读取测试用例数据
def read_data(filename,sheetname):
    wb = load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 选择表单
    cases = [] # 空列表- 存储所有数据库
    max_r = sheet.max_row  # 获取最大行号
    for i in range(2,max_r+1):
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value, # 获取url地址
        data = sheet.cell(row=i,column=6).value, # 获取参数地址
        expected_result = sheet.cell(row=i,column=7).value) # 获取url地址
        cases.append(case)
    return cases    # 返回所有用例

# 写入结果的函数
def write_result(filename,sheetname,row,column,final_result):
    wb = load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 选择表单
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)

"""
{'case_id': 1, 
'url': 'http://47.115.15.198:7001/smarthome/user/register', 
'data': '{"phone":"17750542236","pwd":"1234567a","rePwd":"1234567a","userName":"柠檬小可爱1","verificationCode":"611203"}', 
'expected_result': '{"code":"0","msg":"操作成功"}'
}
real_result = {'code': '1008', 'msg': '注册失败，手机号码已被使用！', 'data': None}
从excel取出来  --字符串

eval(): python内置函数 == 被引号包裹的-Python的表达式取出来 ==  去掉Python表达式外面的引号
"""
def execute(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")   # 获取url
        data = case["data"]   # 获取参数  --字符串-->字典
        data = eval(data)  # 字符串 --> 字典
        expected_result = case.get("expected_result") # 获取期望结果
        expected_result = eval(expected_result) # 字符串 --> 字典
        real_result = api_request(url=url,data=data)   # 调用 发送接口请求函数--传参
        real_msg = real_result.get("msg")   # 执行结果的 msg值
        expected_msg = expected_result["msg"] # 预期结果的 msg值
        print("执行结果：{}".format(real_msg))
        print("期望结果：{}".format(expected_msg))
        if real_msg == expected_msg:
            print("第{}条测试用例执行通过！".format(case_id))
            final = "Passed"
        else:
            print("第{}条测试用例执行不通过！".format(case_id))
            final = "Failed"
        print("*" * 20)
        write_result(filename,sheetname,case_id+1,8,final)  #调用写入结果函数--传参

execute("testcase_api_wuye.xlsx","login")



















